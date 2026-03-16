from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any
from datetime import datetime


class AuditExcelGenerator:
    """Generate professional Excel audit reports."""

    def __init__(self, audit_results: Dict[str, Any]):
        self.audit_results = audit_results
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        # Color scheme
        self.dark_blue = "2C3E50"
        self.light_gray = "ECF0F1"
        self.red_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        self.orange_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        self.green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        self.gray_fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")

        self.white_font = Font(name='Arial', size=11, bold=False, color="FFFFFF")
        self.header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        self.summary_font = Font(name='Arial', size=10, bold=True)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _get_evaluation_fill(self, evaluation: str):
        """Return fill color based on evaluation."""
        if evaluation == "Bad":
            return self.red_fill
        elif evaluation == "Can be Improved":
            return self.orange_fill
        elif evaluation == "Good":
            return self.green_fill
        else:  # N/A
            return self.gray_fill

    def _calculate_scores(self, findings: List[Dict]) -> Dict[str, Any]:
        """Calculate total and percentage scores."""
        total_score = sum(f['score'] for f in findings)
        max_score = sum(f['max_score'] for f in findings)
        percentage = (total_score / max_score * 100) if max_score > 0 else 0

        # Grade assignment
        if percentage >= 90:
            grade = "A"
        elif percentage >= 80:
            grade = "B"
        elif percentage >= 70:
            grade = "C"
        elif percentage >= 60:
            grade = "D"
        else:
            grade = "F"

        return {
            'total_score': total_score,
            'max_score': max_score,
            'percentage': percentage,
            'grade': grade
        }

    def _add_header_row(self, ws, row_num, columns):
        """Add a header row with dark blue background."""
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = self.header_font
            cell.fill = PatternFill(start_color=self.dark_blue, end_color=self.dark_blue, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

    def _create_overview_sheet(self):
        """Create the Overview sheet."""
        ws = self.wb.create_sheet("Overview", 0)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

        # Title
        ws['A1'] = "Website Audit Report"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color="2C3E50")

        # URL and Audit Date
        ws['A3'] = "Audited URL:"
        ws['A3'].font = self.summary_font
        ws['B3'] = self.audit_results.get('url', 'N/A')

        ws['A4'] = "Audit Date:"
        ws['A4'].font = self.summary_font
        ws['B4'] = self.audit_results.get('timestamp', 'N/A')[:10]

        # Score Summary Table
        ws['A6'] = "Score Summary"
        ws['A6'].font = Font(name='Arial', size=12, bold=True, color="2C3E50")

        # Headers for score table
        headers = ['Category', 'Score', 'Max Score', 'Percentage', 'Grade']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = PatternFill(start_color=self.dark_blue, end_color=self.dark_blue, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Calculate scores for each category
        categories = [
            ('SEO', self.audit_results.get('seo', [])),
            ('Core Web Vitals', self.audit_results.get('cwv', [])),
            ('UX & Usability', self.audit_results.get('ux', [])),
            ('Conversion', self.audit_results.get('conversion', []))
        ]

        row = 8
        all_findings = []
        for cat_name, findings in categories:
            if findings:
                all_findings.extend(findings)
                scores = self._calculate_scores(findings)

                ws.cell(row=row, column=1).value = cat_name
                ws.cell(row=row, column=2).value = scores['total_score']
                ws.cell(row=row, column=3).value = scores['max_score']
                ws.cell(row=row, column=4).value = f"{scores['percentage']:.1f}%"
                ws.cell(row=row, column=5).value = scores['grade']

                # Format cells
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(name='Arial', size=10)
                    cell.border = self.border
                    if col >= 2:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                row += 1

        # Overall score
        overall_scores = self._calculate_scores(all_findings)
        row += 1
        ws.cell(row=row, column=1).value = "OVERALL"
        ws.cell(row=row, column=1).font = self.summary_font
        ws.cell(row=row, column=2).value = overall_scores['total_score']
        ws.cell(row=row, column=2).font = self.summary_font
        ws.cell(row=row, column=3).value = overall_scores['max_score']
        ws.cell(row=row, column=3).font = self.summary_font
        ws.cell(row=row, column=4).value = f"{overall_scores['percentage']:.1f}%"
        ws.cell(row=row, column=4).font = self.summary_font
        ws.cell(row=row, column=5).value = overall_scores['grade']
        ws.cell(row=row, column=5).font = self.summary_font

        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill(start_color=self.light_gray, end_color=self.light_gray, fill_type="solid")
            ws.cell(row=row, column=col).border = self.border

        # Identify critical issues (Bad evaluations)
        row += 3
        ws.cell(row=row, column=1).value = "Critical Issues (Score: Bad)"
        ws.cell(row=row, column=1).font = Font(name='Arial', size=11, bold=True, color="E74C3C")

        row += 1
        critical = [f for f in all_findings if f['evaluation'] == 'Bad']
        for issue in critical[:5]:  # Top 5
            ws.cell(row=row, column=1).value = issue['parameter']
            ws.cell(row=row, column=2).value = issue['remarks']
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

        # Quick wins (Good evaluations that could be improved)
        row += 2
        ws.cell(row=row, column=1).value = "Quick Wins (Easy Improvements)"
        ws.cell(row=row, column=1).font = Font(name='Arial', size=11, bold=True, color="27AE60")

        row += 1
        improvements = [f for f in all_findings if f['evaluation'] == 'Can be Improved' and f['impact'] in ['High', 'Medium']]
        for improvement in improvements[:5]:  # Top 5
            ws.cell(row=row, column=1).value = improvement['parameter']
            ws.cell(row=row, column=2).value = improvement['remarks']
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

    def _create_category_sheet(self, sheet_name: str, findings: List[Dict]):
        """Create a category audit sheet (SEO, CWV, UX, Conversion)."""
        ws = self.wb.create_sheet(sheet_name)

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 50

        # Header row
        headers = ['#', 'Parameter', 'Category/Sub-group', 'Evaluation', 'Score', 'Max Score', 'Impact', 'Remarks']
        self._add_header_row(ws, 1, headers)

        # Data rows
        row = 2
        for idx, finding in enumerate(findings, 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = finding['parameter']
            ws.cell(row=row, column=3).value = finding['category']
            ws.cell(row=row, column=4).value = finding['evaluation']
            ws.cell(row=row, column=5).value = finding['score']
            ws.cell(row=row, column=6).value = finding['max_score']
            ws.cell(row=row, column=7).value = finding['impact']
            ws.cell(row=row, column=8).value = finding['remarks']

            # Format cells
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name='Arial', size=10)
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Color-code evaluation column
            eval_cell = ws.cell(row=row, column=4)
            eval_cell.fill = self._get_evaluation_fill(finding['evaluation'])
            eval_cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
            eval_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Center numeric columns
            for col in [1, 5, 6]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        # Summary row
        scores = self._calculate_scores(findings)
        summary_row = row + 1
        ws.cell(row=summary_row, column=2).value = "TOTAL"
        ws.cell(row=summary_row, column=2).font = self.summary_font
        ws.cell(row=summary_row, column=5).value = scores['total_score']
        ws.cell(row=summary_row, column=5).font = self.summary_font
        ws.cell(row=summary_row, column=6).value = scores['max_score']
        ws.cell(row=summary_row, column=6).font = self.summary_font
        ws.cell(row=summary_row, column=8).value = f"{scores['percentage']:.1f}% - Grade: {scores['grade']}"
        ws.cell(row=summary_row, column=8).font = self.summary_font

        for col in range(1, 9):
            cell = ws.cell(row=summary_row, column=col)
            cell.fill = PatternFill(start_color=self.light_gray, end_color=self.light_gray, fill_type="solid")
            cell.border = self.border

        # Freeze panes
        ws.freeze_panes = 'A2'

    def generate(self) -> bytes:
        """Generate the Excel file and return as bytes."""
        # Create Overview sheet
        self._create_overview_sheet()

        # Create category sheets
        self._create_category_sheet("SEO", self.audit_results.get('seo', []))
        self._create_category_sheet("Core Web Vitals", self.audit_results.get('cwv', []))
        self._create_category_sheet("UX & Usability", self.audit_results.get('ux', []))
        self._create_category_sheet("Conversion", self.audit_results.get('conversion', []))

        # Write to BytesIO buffer
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

